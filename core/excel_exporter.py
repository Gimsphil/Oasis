# -*- coding: utf-8 -*-
"""
엑셀 내보내기 모듈 (Excel Exporter)
================================
산출/견적 데이터를 엑셀 파일로 내보내기

기능:
- 총괄표 시트 생성
- 공종별 산출내역서 시트
- 산출일위대 포함
- 서식 적용 (헤더, 테두리, 숫자 형식)
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[WARN] openpyxl 미설치: pip install openpyxl")


def export_to_excel(
    project_data: dict,
    output_path: str,
    include_unit_price: bool = True,
    sheet_name: str = "산출 총괄표",
):
    """
    산출 데이터를 엑셀로 내보내기

    Args:
        project_data: {"공종명": [...(산출 행 데이터)]}
        output_path: 저장 경로 (.xlsx)
        include_unit_price: 산출일위대 포함 여부
        sheet_name: 총괄표 시트명

    Returns:
        bool: 성공 여부
    """
    if not OPENPYXL_AVAILABLE:
        print("[ERROR] openpyxl이 설치되지 않았습니다.")
        return False

    try:
        wb = Workbook()

        # 스타일 정의
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        num_alignment = Alignment(horizontal="right", vertical="center")
        text_alignment = Alignment(horizontal="left", vertical="center")

        # 시트 1: 산출 총괄표
        ws_summary = wb.active
        ws_summary.title = sheet_name[:31]  # 시트명 31자 제한

        # 총괄표 헤더
        summary_headers = ["번호", "공종명", "수량", "비고"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = header_alignment

        # 총괄표 데이터
        for row_idx, (gongjong, items) in enumerate(project_data.items(), 2):
            ws_summary.cell(row=row_idx, column=1, value=row_idx - 1)
            ws_summary.cell(row=row_idx, column=2, value=gongjong)

            # 수량 합산
            total_qty = 0
            for item in items:
                total = item.get("total", "").strip()
                if total:
                    try:
                        total_qty += float(total)
                    except:
                        pass
            ws_summary.cell(row=row_idx, column=3, value=total_qty)

        # 열 너비 조정
        ws_summary.column_dimensions["A"].width = 8
        ws_summary.column_dimensions["B"].width = 40
        ws_summary.column_dimensions["C"].width = 12
        ws_summary.column_dimensions["D"].width = 30

        # 시트 2~N: 공종별 산출내역서
        for gongjong, items in project_data.items():
            sheet_name = gongjong[:31]
            ws = wb.create_sheet(title=sheet_name)

            # 산출내역서 헤더
            eulji_headers = [
                "#.",
                "구분",
                "FROM",
                "TO",
                "회로",
                "산출목록",
                "산출수식",
                "계",
                "단위",
                "비고",
            ]
            for col, header in enumerate(eulji_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = header_alignment

            # 데이터 행
            for row_idx, item in enumerate(items, 2):
                # 번호
                cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
                cell.border = thin_border
                cell.alignment = num_alignment

                # 구분
                cell = ws.cell(row=row_idx, column=2, value=item.get("gubun", ""))
                cell.border = thin_border
                cell.alignment = text_alignment

                # FROM
                cell = ws.cell(row=row_idx, column=3, value=item.get("from", ""))
                cell.border = thin_border

                # TO
                cell = ws.cell(row=row_idx, column=4, value=item.get("to", ""))
                cell.border = thin_border

                # 회로
                cell = ws.cell(row=row_idx, column=5, value=item.get("circuit", ""))
                cell.border = thin_border

                # 산출목록
                cell = ws.cell(row=row_idx, column=6, value=item.get("item", ""))
                cell.border = thin_border
                cell.alignment = text_alignment

                # 산출수식
                cell = ws.cell(row=row_idx, column=7, value=item.get("formula", ""))
                cell.border = thin_border

                # 계
                total = item.get("total", "").strip()
                cell = ws.cell(
                    row=row_idx,
                    column=8,
                    value=float(total)
                    if total and total.replace(".", "").replace("-", "").isdigit()
                    else "",
                )
                cell.border = thin_border
                cell.alignment = num_alignment

                # 단위
                cell = ws.cell(row=row_idx, column=9, value=item.get("unit", ""))
                cell.border = thin_border
                cell.alignment = num_alignment

                # 비고
                cell = ws.cell(row=row_idx, column=10, value=item.get("remark", ""))
                cell.border = thin_border

                # 산출일위대 추가 (있는 경우)
                if include_unit_price and item.get("unit_price"):
                    # 빈 행 추가
                    ws.cell(row=row_idx + 1, column=1, value="")
                    ws.merge_cells(
                        start_row=row_idx + 1,
                        start_column=2,
                        end_row=row_idx + 1,
                        end_column=10,
                    )
                    unit_cell = ws.cell(
                        row=row_idx + 1, column=2, value="  └─ 산출일위대가:"
                    )
                    unit_cell.fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )

                    # 일위대 세부항목
                    unit_items = item.get("unit_price", {}).get("items", [])
                    for u_idx, unit_item in enumerate(unit_items, 1):
                        u_row = row_idx + 1 + u_idx
                        ws.cell(row=u_row, column=2, value=f"    {u_idx}.")
                        ws.cell(row=u_row, column=3, value=unit_item.get("name", ""))
                        ws.cell(row=u_row, column=4, value=unit_item.get("qty", ""))

        # 저장
        wb.save(output_path)
        print(f"[INFO] 엑셀 저장 완료: {output_path}")
        return True

    except Exception as e:
        print(f"[ERROR] 엑셀 내보내기 실패: {e}")
        return False


def export_estimate_to_excel(estimate_data: dict, output_path: str):
    """
    견적 데이터를 엑셀로 내보내기

    Args:
        estimate_data: 견적 데이터 {"공종명": [...]}
        output_path: 저장 경로

    Returns:
        bool: 성공 여부
    """
    if not OPENPYXL_AVAILABLE:
        print("[ERROR] openpyxl이 설치되지 않았습니다.")
        return False

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

        wb = Workbook()

        # 스타일
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 견적서 헤더
        headers = [
            "품명",
            "규격",
            "단위",
            "산출수량",
            "결정수량",
            "단가",
            "금액",
            "구분",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        ws = wb.active
        ws.title = "견적서"

        # 데이터
        for gongjong, items in estimate_data.items():
            # 공종명 행
            ws.cell(row=ws.max_row + 1, column=1, value=gongjong)
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

            for item in items:
                row = ws.max_row + 1
                ws.cell(row=row, column=1, value=item.get("품명", ""))
                ws.cell(row=row, column=2, value=item.get("규격", ""))
                ws.cell(row=row, column=3, value=item.get("단위", ""))
                ws.cell(row=row, column=4, value=item.get("산출수량", ""))
                ws.cell(row=row, column=5, value=item.get("결정수량", ""))
                ws.cell(row=row, column=6, value=item.get("단가", ""))
                ws.cell(row=row, column=7, value=item.get("금액", ""))
                ws.cell(row=row, column=8, value=item.get("구분", ""))

                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = thin_border

        wb.save(output_path)
        print(f"[INFO] 견적서 엑셀 저장 완료: {output_path}")
        return True

    except Exception as e:
        print(f"[ERROR] 견적서 엑셀 내보내기 실패: {e}")
        return False


# ============== 테스트 ==============
if __name__ == "__main__":
    # 테스트 데이터
    test_data = {
        "1. 전등공사": [
            {
                "item": "조명기구 TYPE-A",
                "formula": "10",
                "total": "10",
                "unit": "개",
                "gubun": "조명",
                "remark": "LED 조명",
            },
            {
                "item": "전선 2.5sq",
                "formula": "100+50",
                "total": "150",
                "unit": "m",
                "gubun": "조명",
            },
        ],
        "2. 전열공사": [
            {
                "item": "콘센트",
                "formula": "20",
                "total": "20",
                "unit": "개",
                "gubun": "전열",
            },
        ],
    }

    # 엑셀 내보내기
    export_to_excel(test_data, "test_output.xlsx", include_unit_price=False)
    print("✅ 테스트 완료: test_output.xlsx")
